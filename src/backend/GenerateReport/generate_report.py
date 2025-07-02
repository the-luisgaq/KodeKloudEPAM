import sys
import pandas as pd
import json
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from azure.storage.blob import BlobServiceClient, ContainerClient
from datetime import datetime

warnings.simplefilter("ignore")  # Ignore openpyxl default style warnings

def convert_to_hours(value):
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value = str(value).strip().lower()
    if 'hour' in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0]) if num else 0.0
    elif 'minute' in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0])/60 if num else 0.0
    return 0.0

def download_blob_to_file(container_name, blob_name, local_path):
    connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
    if not connection_string:
        raise Exception("Missing AZURE_STORAGE_CONNECTION_STRING environment variable")

    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    with open(local_path, "wb") as file:
        download_stream = blob_client.download_blob()
        file.write(download_stream.readall())

def upload_file_to_blob(container_name, blob_name, local_path):
    connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
    if not connection_string:
        raise Exception("Missing AZURE_STORAGE_CONNECTION_STRING environment variable")

    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    container_client = blob_service_client.get_container_client(container_name)
    blob_client = container_client.get_blob_client(blob_name)

    # Si el archivo existe, renómbralo con timestamp como backup
    if blob_client.exists():
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        backup_blob_name = f"{blob_name.replace('.json', '')}_{timestamp}.json"
        backup_blob_client = container_client.get_blob_client(backup_blob_name)

        # Copiar el blob actual al de backup
        source_url = blob_client.url
        backup_blob_client.start_copy_from_url(source_url)
        print(f"📦 Created backup version: {backup_blob_name}")

        # Limitar a 5 versiones más recientes
        all_blobs = sorted(
            [b for b in container_client.list_blobs(name_starts_with=blob_name.replace('.json', '')) if b.name.endswith('.json')],
            key=lambda b: b.last_modified,
            reverse=True
        )
        for old_blob in all_blobs[5:]:
            container_client.delete_blob(old_blob.name)
            print(f"🗑️ Deleted old backup: {old_blob.name}")

    # Subir el nuevo archivo
    with open(local_path, "rb") as data:
        blob_client.upload_blob(data, overwrite=True)
        print(f"✅ Uploaded latest: {blob_name}")


def generate_report(admin_path, activity_path, output_excel_path, output_json_path):
    admin_df = pd.read_excel(admin_path)
    activity_df = pd.read_excel(activity_path)

    admin_df['Email'] = admin_df['Email'].str.strip().str.lower()
    activity_df['Email'] = activity_df['Email'].str.strip().str.lower()

    admin_df = admin_df[['Name', 'Email', 'Program', 'License Accepted']]
    activity_df = activity_df[['Email', 'Lessons Completed', 'Video Hours Watched', 'Labs Completed']]

    admin_df = admin_df[admin_df['Program'].str.strip().str.upper() != 'LPC']

    merged = pd.merge(admin_df, activity_df, on='Email', how='left')

    merged['Lessons Completed'] = merged['Lessons Completed'].fillna(0).astype(int)
    merged['Video Hours Watched'] = merged['Video Hours Watched'].apply(convert_to_hours)
    merged['Labs Completed'] = merged['Labs Completed'].fillna(0).astype(int)

    def activity_status(row):
        if row['Lessons Completed'] == 0 and row['Video Hours Watched'] == 0 and row['Labs Completed'] == 0:
            return 'No activity or progress'
        return ''

    merged['Status'] = merged.apply(activity_status, axis=1)

    def license_display(val):
        if isinstance(val, str) and val.strip().lower() == 'no':
            return 'X'
        return '✓'

    merged['License Accepted Display'] = merged['License Accepted'].apply(license_display)

    final_columns = ['Name', 'Email', 'Program', 'Lessons Completed', 'Video Hours Watched', 'Labs Completed', 'License Accepted Display', 'Status']
    display_df = merged[final_columns].rename(columns={"License Accepted Display": "License Accepted"})
    display_df.to_excel(output_excel_path, index=False)

    wb = load_workbook(output_excel_path)
    ws = wb.active

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')

    for row in ws.iter_rows(min_row=2):
        status = row[7].value
        license_col = row[6].value
        if status == 'No activity or progress':
            for cell in row:
                cell.fill = red_fill
        elif license_col == 'X':
            for cell in row:
                cell.fill = orange_fill

    wb.save(output_excel_path)

    json_data = display_df.to_dict(orient='records')
    with open(output_json_path, 'w') as f:
        json.dump(json_data, f, indent=2)

    upload_file_to_blob("kodekloud-inputs", os.path.basename(output_json_path), output_json_path)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_report.py blob:admin.xlsx blob:activity.xlsx")
        sys.exit(1)

    admin_xlsx = sys.argv[1]
    activity_xlsx = sys.argv[2]

    if admin_xlsx.startswith("blob:"):
        blob_name = admin_xlsx[5:]
        download_blob_to_file("kodekloud-inputs", blob_name, "admin.xlsx")
        admin_xlsx = "admin.xlsx"

    if activity_xlsx.startswith("blob:"):
        blob_name = activity_xlsx[5:]
        download_blob_to_file("kodekloud-inputs", blob_name, "activity.xlsx")
        activity_xlsx = "activity.xlsx"

    generate_report(
        admin_xlsx,
        activity_xlsx,
        "kodekloud_report.xlsx",
        "kodekloud_data.json"
    )
